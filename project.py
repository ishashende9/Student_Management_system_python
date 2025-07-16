from openpyxl import Workbook, load_workbook
import os

filename = "students.xlsx"

# Create the Excel file with headers if it doesn't exist
def initialize_excel():
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Contact", "Department"])
        wb.save(filename)
        print(f"✅ Excel file created: {os.path.abspath(filename)}")

# Add student and save to Excel
def add_student():
    name = input("Name: ")
    email = input("Email: ")
    contact = input("Contact: ")
    department = input("Department: ")

    try:
        wb = load_workbook(filename)
        ws = wb.active
        ws.append([name, email, contact, department])
        wb.save(filename)
        print("✅ Student saved!")

        # Open the Excel file after saving
        print("📂 Opening Excel file...")
        os.startfile(filename)

    except PermissionError:
        print("❌ Please close 'students.xlsx' before saving.")

# View student names only
def view_students():
    wb = load_workbook(filename)
    ws = wb.active
    print("\n📋 Student List:")
    count = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # skip empty rows
            continue
        print(f"{count}. {row[0]}")
        count += 1
    print()

# Show details of a specific student
def show_student_details(name):
    wb = load_workbook(filename)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if row[0].lower() == name.lower():
            print(f"\n🔍 Details of {row[0]}:")
            print(f"📧 Email     : {row[1]}")
            print(f"📱 Contact   : {row[2]}")
            print(f"🏫 Department: {row[3]}\n")
            found = True
            break
    if not found:
        print("❌ Student not found.\n")

# Search function
def search_student():
    name = input("Enter name to search: ")
    show_student_details(name)

# Main menu
def menu():
    initialize_excel()
    while True:
        print("\n🎓 Student Management System 🎓")
        print("1️⃣ Add Student")
        print("2️⃣ View Students")
        print("3️⃣ Search Student")
        print("4️⃣ Exit")
        choice = input("Choose (1-4): ")

        if choice == '1':
            add_student()
        elif choice == '2':
            view_students()
            name = input("🔎 Enter name to view details: ")
            show_student_details(name)
        elif choice == '3':
            search_student()
        elif choice == '4':
            print("👋 Exiting... Thank you!")
            break
        else:
            print("❌ Invalid input. Try again.")

# Start the program
menu()

        
